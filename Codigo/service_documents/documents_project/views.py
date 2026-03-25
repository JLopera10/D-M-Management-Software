import json
import logging
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
import openpyxl

logger = logging.getLogger(__name__)

@csrf_exempt
@require_POST
def parse_business_quote(request):
    if 'quote_file' not in request.FILES:
        return JsonResponse({"exito": False, "mensaje": "No file uploaded."}, status=400)
    
    excel_file = request.FILES['quote_file']
    
    if not excel_file.name.endswith('.xlsx'):
        return JsonResponse({"exito": False, "mensaje": "File must be a .xlsx format."}, status=400)

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
        
        obra_raw = str(sheet['A2'].value or "").strip()
        desc_raw = str(sheet['A3'].value or "").strip()
        medidas_raw = str(sheet['A4'].value or "").strip()
        
        nombre_proyecto = obra_raw.replace("OBRA: ", "").strip()
        categoria = desc_raw.replace("DESCRIPCIÓN: ", "").strip()
        medidas = medidas_raw.replace("Medida tramo de ", "").strip()
        fecha_cotizacion = str(sheet['A5'].value or "").strip()

        materiales = []
        mano_de_obra = []
        
        seccion_actual = None
        subtotal = 0.0
        utilidad_factor = ""
        utilidad_valor = 0.0
        total = 0.0

        # Loop through every row starting from row 6 down to the bottom of the sheet
        for row in range(6, sheet.max_row + 1):
            # Read columns A through E (Ignoring the duplicate data on the right)
            col_a = str(sheet.cell(row=row, column=1).value or "").strip()
            col_b = sheet.cell(row=row, column=2).value # Cantidad
            col_c = sheet.cell(row=row, column=3).value # Unidad
            col_d = sheet.cell(row=row, column=4).value # Valor Unitario
            col_e = sheet.cell(row=row, column=5).value # Valor Total

            if not col_a:
                continue # Skip completely empty rows

            # Section Headers
            if "MATERIALES" in col_a.upper():
                seccion_actual = "MATERIALES"
                continue
            elif "MANO DE OBRA" in col_a.upper() and not col_b:
                seccion_actual = "MANO_DE_OBRA"
                continue
                
            # Financial Summaries
            elif "SUBTOTAL" in col_a.upper():
                subtotal = float(col_e or 0.0)
                seccion_actual = None # Stop collecting items
            elif "FACTOR DE UTILIDAD" in col_a.upper():
                utilidad_factor = str(col_d or "").strip()
                utilidad_valor = float(col_e or 0.0)
            elif col_a.upper() == "TOTALES":
                total = float(col_e or 0.0)
                
            # Actual data rows    
            else:
                if seccion_actual and col_e:
                    item = {
                        "descripcion": col_a,
                        "cantidad": float(col_b or 0.0),
                        "unidad": str(col_c or "").strip(),
                        "valor_unitario": float(col_d or 0.0),
                        "valor_total": float(col_e or 0.0)
                    }
                    if seccion_actual == "MATERIALES":
                        materiales.append(item)
                    elif seccion_actual == "MANO_DE_OBRA":
                        mano_de_obra.append(item)

        project_data = {
            "nombre_proyecto": nombre_proyecto,
            "categoria": categoria,
            "medidas": medidas,
            "fecha": fecha_cotizacion,
            "desglose": {
                "materiales": materiales,
                "mano_de_obra": mano_de_obra
            },
            "finanzas": {
                "subtotal": subtotal,
                "utilidad_factor": utilidad_factor,
                "utilidad_valor": utilidad_valor,
                "total": total
            }
        }

        return JsonResponse({
            "exito": True,
            "mensaje": "Cotización virtualizada con éxito.",
            "datos": project_data
        })

    except Exception as e:
        logger.exception("Error procesando Excel:")
        return JsonResponse({
            "exito": False, 
            "mensaje": f"Error interno al procesar el archivo: {str(e)}"
        }, status=500)